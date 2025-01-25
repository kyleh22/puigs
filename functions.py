import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment

def find_combinations_export(df, target, output_file="output_combinations_optimized.xlsx"):
    # Set the first row as column headers
    original_headings = df.iloc[0]  # Store the first row for later re-use
    df = df.rename(columns=dict(zip(df.columns, original_headings)))

    # Remove the row used for headings and reset the index
    df = df[0:].reset_index(drop=True)
    
    # Convert the "Qty" column to numeric
    df.iloc[:, 2] = pd.to_numeric(df.iloc[:, 2], errors="coerce")

    # Extract numeric values from the "Leadtime Weeks" column and add a new column for numeric lead times
    leadtime_column = "Leadtime Weeks"
    df[leadtime_column] = df[leadtime_column].str.extract(r"(\d+)").astype(float)

    # Sort by Leadtime Weeks (ascending) and Qty (descending)
    df.sort_values(by=[leadtime_column, df.columns[2]], ascending=[True, False], inplace=True)

    remaining_df = df.copy()
    group_count = 1
    group_dfs = []

    # Handle items with Qty matching the target
    exact_matches = remaining_df[remaining_df.iloc[:, 2] == target]
    for index, row in exact_matches.iterrows():
        group_df = pd.DataFrame([row], columns=remaining_df.columns)
        group_df["Group"] = f"Container {group_count}"

        # Add a total row for the group
        total_row = {col: None for col in group_df.columns}
        total_row[df.columns[2]] = target
        total_row["Group"] = f"Total for Container {group_count}"
        group_df = pd.concat([group_df, pd.DataFrame([total_row])], ignore_index=True)

        group_dfs.append(group_df)
        group_count += 1

    # Remove exact matches from the remaining dataframe
    remaining_df = remaining_df[remaining_df.iloc[:, 2] != target]

    # Process the rest of the items
    while not remaining_df.empty:
        current_sum = 0
        current_group = []
        indexes_to_remove = []

        for index, row in remaining_df.iterrows():
            qty = row.iloc[2]

            if current_sum + qty <= target:
                current_sum += qty
                current_group.append(row)
                indexes_to_remove.append(index)
            elif current_sum < target:
                # Split the row to fit the remaining capacity
                split_amount = target - current_sum
                split_row = row.copy()
                split_row.iloc[2] = split_amount
                current_group.append(split_row)

                # Update the remaining quantity in the original row
                remaining_df.at[index, remaining_df.columns[2]] -= split_amount
                break

        # Remove used items from remaining_df
        remaining_df.drop(index=indexes_to_remove, inplace=True)

        # Create a DataFrame for the current group
        group_df = pd.DataFrame(current_group, columns=remaining_df.columns)
        group_df["Group"] = f"Container {group_count}"

        # Add a total row for the group
        total_row = {col: None for col in group_df.columns}
        total_row[df.columns[2]] = group_df.iloc[:, 2].sum()
        total_row["Group"] = f"Total for Container {group_count}"
        group_df = pd.concat([group_df, pd.DataFrame([total_row])], ignore_index=True)

        group_dfs.append(group_df)

        # Re-sort remaining_df to maintain lead time priority
        remaining_df = remaining_df[remaining_df.iloc[:, 2] > 0]  # Remove rows with zero quantity
        remaining_df.sort_values(by=[leadtime_column, remaining_df.columns[2]], ascending=[True, False], inplace=True)

        group_count += 1

    # Handle the last container without worrying about the target
    if not remaining_df.empty:
        final_group = remaining_df.copy()
        final_group["Group"] = f"Container {group_count}"

        # Add a total row for the final group
        total_row = {col: None for col in final_group.columns}
        total_row[df.columns[2]] = final_group.iloc[:, 2].sum()
        total_row["Group"] = f"Total for Container {group_count}"
        final_group = pd.concat([final_group, pd.DataFrame([total_row])], ignore_index=True)

        group_dfs.append(final_group)

    # Handle cases where the sum is less than the target
    for group_df in group_dfs:
        total_qty = group_df.iloc[:-1, 2].sum()  # Exclude the total row
        if total_qty < target:
            total_row = group_df.iloc[-1].copy()
            total_row[df.columns[2]] = total_qty
            total_row["Group"] = f"Total for Container {group_count}"
            group_df.iloc[-1] = total_row

    # Combine all groups into a single DataFrame
    formatted_dfs = []
    for i, group_df in enumerate(group_dfs):
        title_row = pd.DataFrame([[None] * (len(group_df.columns) - 1) + [f"Container {i + 1}"]], columns=group_df.columns)
        formatted_dfs.extend([title_row, group_df, pd.DataFrame(columns=group_df.columns)])

    combined_df = pd.concat(formatted_dfs, ignore_index=True)

    return combined_df


def clean_group_column(df):
    # Replace 'Group' column values with an empty string where 'Qty' is None
    df.loc[df['Qty'].isna(), 'Group'] = ''
    return df

def export_df_to_excel(df, input_file, target):
    # Export to Excel
    str_target = str(target)
    str_input = str(input_file)
    output_file = str_input+"_"+str_target+'.xlsx'
    df.to_excel(output_file, index=False, sheet_name="All Groups", engine="openpyxl")

    # Style the Excel file
    wb = load_workbook(output_file)
    ws = wb["All Groups"]

    # Define styles
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Find "Qty" column
    qty_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):  # Assuming the first row contains headers
        if cell.value == "Qty":
            qty_col_idx = idx
        if qty_col_idx:
            break

    if qty_col_idx is None:
        print("Could not find 'Qty' column in the Excel file.")
        return

    # Apply styles
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        is_total_row = row[1].value and "Total for Container" in str(row[1].value)
        is_title_row = row[1].value and "Container" in str(row[1].value) and "Total" not in str(row[1].value)

        for idx, cell in enumerate(row, start=1):
            if idx == qty_col_idx:  # Only highlight the Qty column
                if is_total_row:
                    cell.fill = highlight_fill
            if cell.value is not None:  # Apply borders to non-empty cells
                cell.border = border_style
            if is_title_row:  # Center-align title rows
                cell.alignment = center_alignment

    # Save the updated workbook
    wb.save(output_file)
    print(f"Results exported to {output_file}")
